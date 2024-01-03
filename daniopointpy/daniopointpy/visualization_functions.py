import os
import zipfile

import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
import io
import seaborn as sns

from statsmodels.stats.multicomp import pairwise_tukeyhsd
from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage


def visualize_plate(file, location, endpoint):
    if format(extension) == ".xlsx":
        df_raw = pd.read_excel(file)
    else:
        df_raw = pd.read_table(file, encoding="utf-16", low_memory=False)
    well_numbers = df_raw[location].nunique()
    df_1 = df_raw[[location, endpoint]]
    df_2 = df_1.groupby(location).sum()
    array = df_2.values
    if well_numbers == 24:
        num2 = array.reshape(4, 6)
        columns = list(range(1, 7))
        rows = [chr(ord('A') + i) for i in range(4)]
    elif well_numbers == 48:
        num2 = array.reshape(6, 8)
        columns = list(range(1, 9))
        rows = [chr(ord('A') + i) for i in range(6)]
    elif well_numbers == 96:
        num2 = array.reshape(8, 12)
        columns = list(range(1, 13))
        rows = [chr(ord('A') + i) for i in range(8)]
    else:
        print("Plate format not supported... Sorry!")
        exit()
    plt.imshow(num2, cmap='Reds', interpolation='nearest')
    plt.xticks(np.arange(len(columns)), columns)
    plt.yticks(np.arange(len(rows)), rows)
    plt.colorbar(label='Values')
    plt.xlabel('Columns')
    plt.ylabel('Rows')
    plt.title('Heatmap Example')
    plt.show()


def well_response(file, location, timepoint, endpoint):
    current_row = 2
    wb = Workbook()
    ws = wb.active
    ws['A1'] = 'Well Name'
    ws['B1'] = 'Image'
    ws['C1'] = f'Sum of {endpoint} along Measurement'
    file_name = os.path.basename(file)
    name, extension = os.path.splitext(file_name)
    if format(extension) == ".xlsx":
        df_raw = pd.read_excel(file)
    else:
        df_raw = pd.read_table(file, encoding="utf-16", low_memory=False)
    well_numbers = df_raw[location].nunique()
    if well_numbers not in [24, 48, 96]:
        raise ValueError("Invalid plate size. Supported plate sizes are 24, 48, and 96.")
    locations_per_row = well_numbers // 8
    location_mapping = {f'c{i:02d}': f'{chr(65 + (i - 1) // locations_per_row)}{((i - 1) % locations_per_row) + 1:02d}'
                        for i in range(1, plate_size + 1)}
    df['well_plate_position'] = df[location].map(location_mapping)
    wells = df['well_plate_position'].unique()

    for well in wells:
        well_data = df[df['well_plate_position'] == well]
        time = well_data[timepoint]
        dist = well_data[endpoint]
        total_endpoint = dist.sum()
        ws[f'C{current_row}'] = total_endpoint

        plt.close()
        plt.plot(time, dist)
        plt.title(f'Well {well}')
        plt.xlabel('Time')
        plt.ylabel('Distance')
        plt.grid(True)

        img_stream = io.BytesIO()
        plt.savefig(img_stream, format='png')
        img_stream.seek(0)

        img = ExcelImage(img_stream)
        img.width = 200  # Adjust the image width as needed
        img.height = 150  # Adjust the image height as needed
        ws.add_image(img, f'B{current_row}')
        row_letter, col_number = well[0], int(well[1:])
        well_name = f'{row_letter}{col_number:02d}'
        ws[f'A{current_row}'] = well_name
        current_row += 1

    for row in ws.iter_rows(min_row=2, max_row=current_row, max_col=2):
        for cell in row:
            cell.alignment = cell.alignment.copy(wrap_text=True)
            col_letter = cell.column_letter
            col_width = ws.column_dimensions[col_letter].width
            ws.column_dimensions[col_letter].width = max(col_width, img.width / 6)  # Adjust the divisor as needed

    excel_file_path = f'test/well_charts_{plate_size}.xlsx'
    wb.save(excel_file_path)

    print(f'Excel file saved to {excel_file_path}')

    zip_file_path = f'test/well_charts_{plate_size}.zip'
    with zipfile.ZipFile(zip_file_path, 'w', zipfile.ZIP_DEFLATED) as zf:
        zf.write(excel_file_path, os.path.basename(excel_file_path))

    print(f'Zip file containing Excel file created at {zip_file_path}')


def visualize_plots(file, location, endpoint, treatments_file=None, save_plots=True, start_range=None):
    if format(extension) == ".xlsx":
        df = pd.read_excel(file)
    else:
        df = pd.read_table(file, encoding="utf-16", low_memory=False)
    if treatments_file is not None:
        treatments_df = pd.read_excel(treatments_file)
        animal_treatments = dict(zip(treatments_df['well_code'], treatments_df['treatment']))
        df['treatment'] = df[location].map(animal_treatments)
    else:
        df['treatment'] = 'Unknown'
    if start_range is not None:
        df = df[(df['start'] >= start_range[0]) & (df['start'] <= start_range[1])]

    plt.figure(figsize=(10, 6))
    sns.boxplot(x='treatment', y=endpoint, data=df, palette='Set3',
                showfliers=False)  # Set showfliers=False to exclude outliers
    plt.title(f'Boxplot of {endpoint} by Treatment within Start Range')
    if save_plots:
        plt.savefig(f'test/boxplot_{endpoint}_start_range.png')
    else:
        plt.show()
    plt.close()

    plt.figure(figsize=(10, 6))
    sns.lineplot(x='start', y=endpoint, hue='treatment', data=df)
    plt.title(f'Time Series Line Plot of {endpoint} by Treatment within Start Range')
    if save_plots:
        plt.savefig(f'test/lineplot_{endpoint}_start_range.png')
    else:
        plt.show()
    plt.close()

    plt.figure(figsize=(10, 6))
    sns.histplot(data=df, x=endpoint, hue='treatment', multiple="stack", kde=True)
    plt.title(f'Distribution of {endpoint} by Treatment within Start Range')
    if save_plots:
        plt.savefig(f'test/histplot_{endpoint}_start_range.png')
    else:
        plt.show()
    plt.close()

    tukey_results = pairwise_tukeyhsd(df[endpoint], df['treatment'])

    plt.figure(figsize=(10, 6))
    sns.boxplot(x='treatment', y=endpoint, data=df, palette='Set3', showfliers=False)
    for i, treatment in enumerate(df['treatment'].unique()):
        p_value = tukey_results.pvalues[i]
        stars = ""
        if p_value < 0.05:
            if p_value < 0.001:
                stars = "***"
            elif p_value < 0.01:
                stars = "**"
            else:
                stars = "*"
        plt.text(i, max(df[endpoint]), stars, ha='center', va='bottom', color='red', fontsize=12)
    plt.title(f'Boxplot of {endpoint} by Treatment within Start Range')

    if save_plots:
        plt.savefig(f'test/boxplot_with_stars_{endpoint}_start_range.png')
    else:
        plt.show()
    plt.close()
