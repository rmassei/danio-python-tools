import os
import zipfile

import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
import io
import seaborn as sns

from statsmodels.stats.multicomp import pairwise_tukeyhsd
from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage


def visualize_plate(file, location, endpoint, plate_type=96):
    file_name = os.path.basename(file)
    name, extension = os.path.splitext(file_name)
    if format(extension) == ".xlsx":
        df_raw = pd.read_excel(file)
    else:
        df_raw = pd.read_table(file, encoding="utf-16", low_memory=False)
    if plate_type not in [24, 48, 96]:
        raise ValueError("Unsupported plate type. Supported types are 24, 48, and 96.")
    num_rows = 8
    num_columns = 12 if plate_type == 96 else 6
    if plate_type == 24:
        location_mapping = {f'c{i:02d}': f'{chr(65 + (i - 1) // 6)}{((i - 1) % 6) + 1:02d}'
                            for i in range(1, 25)}
    elif plate_type == 48:
        location_mapping = {f'c{i:02d}': f'{chr(65 + (i - 1) // 6)}{((i - 1) % 6) + 1:02d}'
                            for i in range(1, 49)}
    else:
        location_mapping = {f'c{i:02d}': f'{chr(65 + (i - 1) // 12)}{((i - 1) % 12) + 1:02d}'
                            for i in range(1, 97)}
    df_raw['well_plate_position'] = df_raw[location].map(location_mapping)
    well_mapping = {f'{chr(65 + r)}{c + 1:02d}': (r, c) for r in range(num_rows)
                    for c in range(num_columns)}
    plate = np.zeros((num_rows, num_columns))
    for _, row in df_raw.iterrows():
        well = row['well_plate_position']
        measurement = row[endpoint]
        row_idx, col_idx = well_mapping.get(well, (-1, -1))
        if row_idx != -1 and col_idx != -1:
            plate[row_idx, col_idx] += measurement
    plt.imshow(plate, cmap='Reds')
    plt.colorbar()
    plt.title(f'Results - Sum of {endpoint}')
    plt.xticks(range(num_columns), range(1, num_columns + 1))
    plt.yticks(range(num_rows), [chr(65 + r) for r in range(num_rows)])
    for well, (row, col) in well_mapping.items():
        plt.text(col, row, well, ha='center', va='center', color='w')
    plt.show()


def well_response(file, location, timepoint, endpoint, plate_size=96):
    current_row = 2
    wb = Workbook()
    ws = wb.active
    ws['A1'] = 'Well Name'
    ws['B1'] = 'Image'
    ws['C1'] = f'Sum of {endpoint} along Measurement'
    file_name = os.path.basename(file)
    name, extension = os.path.splitext(file_name)
    if format(extension) == ".xlsx":
        df = pd.read_excel(file)
    else:
        df = pd.read_table(file, encoding="utf-16", low_memory=False)
    if plate_size not in [24, 48, 96]:
        raise ValueError("Invalid plate size. Supported plate sizes are 24, 48, and 96.")
    locations_per_row = plate_size // 8
    location_mapping = {f'c{i:02d}': f'{chr(65 + (i - 1) // locations_per_row)}{((i - 1) % locations_per_row) + 1:02d}'
                        for i in range(1, plate_size + 1)}
    df['well_plate_position'] = df[location].map(location_mapping)
    wells = df['well_plate_position'].unique()

    # Add Well number to the "Well Name" column
    for well in wells:
        well_data = df[df['well_plate_position'] == well]
        time = well_data[timepoint]
        dist = well_data[endpoint]

        # Calculate total endpoint and add to Excel
        total_endpoint = dist.sum()
        ws[f'C{current_row}'] = total_endpoint

        plt.close()
        plt.plot(time, dist)
        plt.title(f'Well {well}')
        plt.xlabel('Time')
        plt.ylabel('Distance')
        plt.grid(True)

        # Save the image in memory
        img_stream = io.BytesIO()
        plt.savefig(img_stream, format='png')
        img_stream.seek(0)

        # Create ExcelImage from the in-memory image
        img = ExcelImage(img_stream)
        img.width = 200  # Adjust the image width as needed
        img.height = 150  # Adjust the image height as needed
        ws.add_image(img, f'B{current_row}')

        # Extract row and column from the well name
        row_letter, col_number = well[0], int(well[1:])

        # Calculate the corresponding well name in the format A01, A02, ..., H11, H12
        well_name = f'{row_letter}{col_number:02d}'

        # Add the well name to the "Well Name" column
        ws[f'A{current_row}'] = well_name

        current_row += 1

    # Adjust cell dimensions based on image size
    for row in ws.iter_rows(min_row=2, max_row=current_row, max_col=2):
        for cell in row:
            cell.alignment = cell.alignment.copy(wrap_text=True)
            col_letter = cell.column_letter
            col_width = ws.column_dimensions[col_letter].width
            ws.column_dimensions[col_letter].width = max(col_width, img.width / 6)  # Adjust the divisor as needed

    # Save the Excel file
    excel_file_path = f'test/well_charts_{plate_size}.xlsx'
    wb.save(excel_file_path)

    print(f'Excel file saved to {excel_file_path}')

    # Create a zip file containing the Excel file
    zip_file_path = f'test/well_charts_{plate_size}.zip'
    with zipfile.ZipFile(zip_file_path, 'w', zipfile.ZIP_DEFLATED) as zf:
        zf.write(excel_file_path, os.path.basename(excel_file_path))

    print(f'Zip file containing Excel file created at {zip_file_path}')


def visualize_plots(file, location, endpoint, plate_type=96, treatments_file=None, save_plots=True, start_range=None):
    if plate_type not in [24, 48, 96]:
        raise ValueError("Unsupported plate type. Supported types are 24, 48, and 96.")
    file_name = os.path.basename(file)
    name, extension = os.path.splitext(file_name)
    if format(extension) == ".xlsx":
        df = pd.read_excel(file)
    else:
        df = pd.read_table(file, encoding="utf-16", low_memory=False)

    if treatments_file is not None:
        treatments_df = pd.read_excel(treatments_file)
        animal_treatments = dict(zip(treatments_df['well_code'], treatments_df['treatment']))
        df['treatment'] = df[location].map(animal_treatments)
    else:
        df['treatment'] = 'Unknown'

    # Filter the DataFrame based on the start_range
    if start_range is not None:
        df = df[(df['start'] >= start_range[0]) & (df['start'] <= start_range[1])]

    # Boxplot by Treatment with different colors
    plt.figure(figsize=(10, 6))
    sns.boxplot(x='treatment', y=endpoint, data=df, palette='Set3',
                showfliers=False)  # Set showfliers=False to exclude outliers
    plt.title(f'Boxplot of {endpoint} by Treatment within Start Range')

    if save_plots:
        plt.savefig(f'test/boxplot_{endpoint}_start_range.png')
    else:
        plt.show()
    plt.close()

    # Time series line plot with 'end' variable on the y-axis
    plt.figure(figsize=(10, 6))
    sns.lineplot(x='start', y=endpoint, hue='treatment', data=df)
    plt.title(f'Time Series Line Plot of {endpoint} by Treatment within Start Range')

    if save_plots:
        plt.savefig(f'test/lineplot_{endpoint}_start_range.png')
    else:
        plt.show()
    plt.close()

    # Distribution plot by Treatment
    plt.figure(figsize=(10, 6))
    sns.histplot(data=df, x=endpoint, hue='treatment', multiple="stack", kde=True)
    plt.title(f'Distribution of {endpoint} by Treatment within Start Range')

    if save_plots:
        plt.savefig(f'test/histplot_{endpoint}_start_range.png')
    else:
        plt.show()
    plt.close()

    # Perform Tukey's HSD test for multiple comparisons
    tukey_results = pairwise_tukeyhsd(df[endpoint], df['treatment'])

    # Boxplot annotation with significance stars
    plt.figure(figsize=(10, 6))
    sns.boxplot(x='treatment', y=endpoint, data=df, palette='Set3', showfliers=False)
    for i, treatment in enumerate(df['treatment'].unique()):
        p_value = tukey_results.pvalues[i]
        stars = ""
        if p_value < 0.05:
            if p_value < 0.001:
                stars = "***"
            elif p_value < 0.01:
                stars = "**"
            else:
                stars = "*"
        plt.text(i, max(df[endpoint]), stars, ha='center', va='bottom', color='red', fontsize=12)
    plt.title(f'Boxplot of {endpoint} by Treatment within Start Range')

    if save_plots:
        plt.savefig(f'test/boxplot_with_stars_{endpoint}_start_range.png')
    else:
        plt.show()
    plt.close()


